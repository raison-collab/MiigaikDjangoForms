import random
from pprint import pprint

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .data.questions import Questions
from .models import AnswerModel, QuestionsModel


class Util:
    @staticmethod
    def shuffle_the_list(lst: list) -> list:
        random.shuffle(lst)
        return lst

    @staticmethod
    def reformat_answers(answers: list) -> list:
        questions = Questions()

        ans_text = []

        for row_index, row in enumerate(answers, start=1):
            row_info = []

            for ans in QuestionsModel().get_fields():
                if ans in ['q1_dop', 'q13', 'q18_dop']:
                    continue

                row_info.append(questions.get_questions_ans(int(ans.replace('q', '')), ans_id=row[ans])[0]['text'])

            row_info.insert(1, row['q1_dop'])
            row_info.insert(13, row['q13'])
            row_info.insert(19, row['q18_dop'])

            ans_text.append(row_info)

        return ans_text

    @staticmethod
    def reformat_asks(asks: list) -> list:
        asks.insert(1, 'Вопрос 1 (если Да)')
        asks.insert(19, 'Вопрос 18 (если Да)')
        return asks

    @staticmethod
    def generate_xlsx_file(path: str,
                           students_data: list[list],
                           students_headers: list,
                           ans_data: list[list],
                           ans_headers: list):
        wb = Workbook()

        wb.create_sheet('ответы', 0)
        wb.create_sheet('студенты', 1)

        ans_ws: Worksheet = wb.get_sheet_by_name('ответы')
        students_ws: Worksheet = wb.get_sheet_by_name('студенты')

        ans_ws.append(ans_headers)
        students_ws.append(students_headers)

        for row in ans_data:
            ans_ws.append(row)

        for row in students_data:
            students_ws.append(row)

        wb.save(path)

